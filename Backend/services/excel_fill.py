from pathlib import Path
from typing import Optional, List, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet
import copy
import re
import unicodedata

TEMPLATE_CONCEPTO = "3_y_4_Concepto_tecnico_y_sectorial_2025.xlsx"
TEMPLATE_CADENA = "6.Cadena_de_valor.xlsx"
TEMPLATE_VIABILIDAD = "7.Viabilidad_dependencias.xlsx"
OUTPUT_CONCEPTO = "{}_3_y_4_Concepto_tecnico_y_sectorial_2025.xlsx"
OUTPUT_CADENA = "{}_6.Cadena_de_valor.xlsx"
OUTPUT_VIABILIDAD = "{}_7.Viabilidad_dependencias.xlsx"

def _norm(s: str) -> str:
        return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c)).lower()

def _get_sheet_fuzzy(wb, target_name: str):
    try:
        return wb[target_name]
    except KeyError:
        pass
    tnorm = _norm(target_name)
    for name in wb.sheetnames:
        if _norm(name) == tnorm or tnorm in _norm(name):
            return wb[name]
    raise KeyError(f"No se encontró la hoja '{target_name}'")

def _next_sequential_index(out_dir: Path) -> int:
    rx = re.compile(r"^(\d+)_3_y_4_Concepto_tecnico_y_sectorial_2025\.xlsx$", re.I)
    max_n = 0
    for p in out_dir.glob("*_3_y_4_Concepto_tecnico_y_sectorial_2025.xlsx"):
        m = rx.match(p.name)
        if m:
            try:
                n = int(m.group(1))
                if n > max_n:
                    max_n = n
            except ValueError:
                pass
    return max_n + 1

def _anchor_of_merged(ws: Worksheet, coord: str) -> str:
    r, c = coordinate_to_tuple(coord)
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
            return ws.cell(row=mr.min_row, column=mr.min_col).coordinate
    return coord

def _write(ws: Worksheet, coord: str, value):
    ws[_anchor_of_merged(ws, coord)] = value

def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int):
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        s = ws.cell(row=src_row, column=col)
        d = ws.cell(row=dst_row, column=col)
        if s.has_style:
            d.font = copy.copy(s.font)
            d.border = copy.copy(s.border)
            d.fill = copy.copy(s.fill)
            d.number_format = s.number_format
            d.protection = copy.copy(s.protection)
            d.alignment = copy.copy(s.alignment)

def _get_block_merges(ws: Worksheet, src_start: int, src_end: int) -> List[Tuple[int, int, int, int]]:
    out: List[Tuple[int, int, int, int]] = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row >= src_start and mr.max_row <= src_end:
            out.append((mr.min_row, mr.min_col, mr.max_row, mr.max_col))
    return out

def _apply_block_merges(ws: Worksheet, merges: List[Tuple[int, int, int, int]], dst_start: int, src_start: int):
    off = dst_start - src_start
    for r1, c1, r2, c2 in merges:
        ws.merge_cells(start_row=r1 + off, start_column=c1, end_row=r2 + off, end_column=c2)

def _clone_block_styles_merges(ws: Worksheet, src_start: int, src_end: int, dst_start: int, merges_cache=None):
    for i in range(src_end - src_start + 1):
        _copy_row_style(ws, src_start + i, dst_start + i)
    merges = merges_cache if merges_cache is not None else _get_block_merges(ws, src_start, src_end)
    _apply_block_merges(ws, merges, dst_start, src_start)

def _find_label_anchor(ws: Worksheet, row: int, text: str):
    for col in range(1, ws.max_column + 1):
        if (ws.cell(row=row, column=col).value or "") == text:
            return (row, col)
    return None

def _find_value_anchor_col_for_row(ws: Worksheet, row: int, label_col: Optional[int]) -> int:
    candidates = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row == row and mr.max_row == row:
            candidates.append(mr)
    candidates.sort(key=lambda r: (r.max_col - r.min_col, r.min_col), reverse=True)
    for mr in candidates:
        if label_col is None or not (mr.min_col <= label_col <= mr.max_col):
            return mr.min_col
    return 3 

def _collect_merges_in_rows(ws: Worksheet, start_row: int) -> List[Tuple[int, int, int, int]]:
    out: List[Tuple[int, int, int, int]] = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row >= start_row:
            out.append((mr.min_row, mr.min_col, mr.max_row, mr.max_col))
    return out

def _unmerge_ranges(ws: Worksheet, ranges: List[Tuple[int, int, int, int]]):
    for r1, c1, r2, c2 in ranges:
        try:
            ws.unmerge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        except Exception:
            pass

def _remerge_with_offset(ws: Worksheet, ranges: List[Tuple[int, int, int, int]], row_off: int):
    for r1, c1, r2, c2 in ranges:
        ws.merge_cells(start_row=r1 + row_off, start_column=c1, end_row=r2 + row_off, end_column=c2)

def _move_down_from_row(ws: Worksheet, start_row: int, row_off: int):
    if row_off <= 0:
        return
    max_col_letter = get_column_letter(ws.max_column)
    max_row = ws.max_row
    heights = {r: ws.row_dimensions[r].height for r in range(start_row, max_row + 1)}
    merges = _collect_merges_in_rows(ws, start_row)
    _unmerge_ranges(ws, merges)
    ws.move_range(f"A{start_row}:{max_col_letter}{max_row}", rows=row_off, cols=0, translate=True)
    for r, h in heights.items():
        ws.row_dimensions[r + row_off].height = h
    _remerge_with_offset(ws, merges, row_off)

def fill_from_template(base_dir: Path, data: dict, force_index: Optional[int] = None, output_dir: Optional[Path] = None) -> Path:
    template_path = base_dir / TEMPLATE_CONCEPTO
    if not template_path.exists():
        raise FileNotFoundError(f"No se encontró el template: {template_path}")

    wb = load_workbook(filename=str(template_path))
    ws = wb.active
    numero_meta: List[str] = list(map(str, data.get("numero_meta", [])))
    nombre_meta: List[str] = list(map(str, data.get("nombre_meta", [])))
    total_metas = max(len(numero_meta), len(nombre_meta))
    extra = max(0, total_metas - 1)
    shift = extra * 3

    if shift > 0:
        _move_down_from_row(ws, start_row=15, row_off=shift)
        merges_base = _get_block_merges(ws, 12, 14)
        for i in range(2, total_metas + 1):
            dst_start = 12 + (i - 1) * 3
            _clone_block_styles_merges(ws, 12, 14, dst_start, merges_cache=merges_base)
            lbl_num = _find_label_anchor(ws, 12, "NÚMERO DE META")
            lbl_nom = _find_label_anchor(ws, 13, "META DE CUATRIENIO")
            if lbl_num:
                _write(ws, f"{get_column_letter(lbl_num[1])}{dst_start}", "NÚMERO DE META")
            if lbl_nom:
                _write(ws, f"{get_column_letter(lbl_nom[1])}{dst_start+1}", "META DE CUATRIENIO")

    _write(ws, "D3",  data.get("nombre_proyecto", ""))
    _write(ws, "C5",  data.get("cod_id_mga", ""))
    _write(ws, "F5",  data.get("nombre_dependencia", ""))
    _write(ws, "C8",  data.get("codigo_sector", ""))
    _write(ws, "F8",  data.get("nombre_sector", ""))
    _write(ws, "C9",  data.get("codigo_programa", ""))
    _write(ws, "F9",  data.get("nombre_programa", ""))
    _write(ws, "C10", data.get("nombre_linea_estrategica", ""))

    # ---------------------------
    # 2) Variables SECTORIALES (Hoja 1)
    # ---------------------------
    variables_sec = data.get("variables_sectorial_respuestas")
    if variables_sec is None:
        raw = data.get("variables_sectorial", data.get("variables", []))
        variables_sec = [_to_respuesta(v) for v in raw]

    for i, base_row in enumerate(range(35, 44)):
        cell = f"H{base_row + shift}"
        v = variables_sec[i] if i < len(variables_sec) else ""
        _write(ws, cell, v)

    # ---------------------------
    # 2.1) Variables TÉCNICAS
    # ---------------------------

    ws_tecnico = _get_sheet_fuzzy(wb, "Concepto Técnico General")
    variables_tec = data.get("variables_tecnico_respuestas")
    if variables_tec is None:
        rawt = data.get("variables_tecnico", [])
        variables_tec = [_to_respuesta(v) for v in rawt]

    for i, base_row in enumerate(range(34, 47)):
        cell = f"H{base_row}"
        v = variables_tec[i] if i < len(variables_tec) else ""
        _write(ws_tecnico, cell, v)

    # ---------------------------
    # Firmas / constancias (ambas hojas)
    # ---------------------------
    fecha_firma = data.get("fecha_firma_texto", "")
    firma_dep   = data.get("firma_secretaria_texto", "")
    _write(ws, f"B{56 + shift}", fecha_firma)
    _write(ws, f"B{58 + shift}", firma_dep)
    _write(ws_tecnico, "B55", fecha_firma)
    _write(ws_tecnico, "B57", firma_dep)

    # ---------------------------
    # 3) Políticas / Categorías / Subcategorías / Valor destinado 
    # ---------------------------
    def _pair(lst, default=""):
        a = lst[0] if len(lst) >= 1 else default
        b = lst[1] if len(lst) >= 2 else default
        return a, b

    nombre_politica = data.get("nombre_politica", []) or data.get("nombre_politica".replace("ó", "o"), [])
    p1, p2 = _pair(list(map(str, nombre_politica)))
    _write(ws, f"E{47 + shift}", p1); _write(ws, f"G{47 + shift}", p2)

    nombre_categoria = data.get("nombre_categoria", [])
    c1, c2 = _pair(list(map(str, nombre_categoria)))
    _write(ws, f"E{48 + shift}", c1); _write(ws, f"G{48 + shift}", c2)

    nombre_focalizacion = data.get("nombre_focalización", []) or data.get("nombre_focalizacion", [])
    f1, f2 = _pair(list(map(str, nombre_focalizacion)))
    _write(ws, f"E{49 + shift}", f1); _write(ws, f"G{49 + shift}", f2)

    valores = data.get("valor_destinado", [])
    v1 = valores[0] if len(valores) > 0 else None
    v2 = valores[1] if len(valores) > 1 else None
    _write(ws, f"E{50 + shift}", v1 if v1 is not None else "")
    _write(ws, f"G{50 + shift}", v2 if v2 is not None else "")

    # ---------------------------
    # 4.5) Estructura financiera
    # ---------------------------
    ef_rows = data.get("estructura_financiera", [])
    if ef_rows:
        years = sorted({row.get("anio") for row in ef_rows if row.get("anio") is not None})[:4]
        while len(years) < 4:
            years.append(None)

        ENT_ORDER = ["PROPIOS", "SGP_LIBRE_INVERSION", "SGP_LIBRE_DESTINACION", "SGP_APSB", "SGP_EDUCACION", "SGP_ALIMENTACION_ESCOLAR", "SGP_CULTURA", "SGP_DEPORTE", "SGP_SALUD", "MUNICIPIO", "NACION", "OTROS"]
        row_by_ent = {"PROPIOS": 19, 
                      "SGP_LIBRE_INVERSION": 20, 
                      "SGP_LIBRE_DESTINACION": 21, 
                      "SGP_APSB": 22, 
                      "SGP_EDUCACION": 23, 
                      "SGP_ALIMENTACION_ESCOLAR": 24, 
                      "SGP_CULTURA": 25, 
                      "SGP_DEPORTE": 26, 
                      "SGP_SALUD": 27,
                      "MUNICIPIO": 28, 
                      "NACION": 29, 
                      "OTROS": 30}
        col_by_idx = {0: "C", 1: "E", 2: "F", 3: "G"}

        from decimal import Decimal
        lookup = {}
        for r in ef_rows:
            anio = r.get("anio")
            ent = (r.get("entidad") or "").strip().upper()
            valor = r.get("valor")
            if ent in row_by_ent:
                lookup[(anio, ent)] = valor if valor is not None else Decimal("0")

        for yi, y in enumerate(years):
            for ent in ENT_ORDER:
                base_row = row_by_ent[ent]
                col = col_by_idx[yi]
                dest = f"{col}{base_row + shift}"
                val = lookup.get((y, ent), Decimal("0"))
                _write(ws, dest, val)
    
    if ef_rows:
        ENT_ORDER = ["PROPIOS", "SGP_LIBRE_INVERSION", "SGP_LIBRE_DESTINACION", "SGP_APSB", "SGP_EDUCACION", "SGP_ALIMENTACION_ESCOLAR", "SGP_CULTURA", "SGP_DEPORTE", "SGP_SALUD", "MUNICIPIO", "NACION", "OTROS"]
        row_by_ent = {"PROPIOS": 18, 
                      "SGP_LIBRE_INVERSION": 19, 
                      "SGP_LIBRE_DESTINACION": 20, 
                      "SGP_APSB": 21, 
                      "SGP_EDUCACION": 22, 
                      "SGP_ALIMENTACION_ESCOLAR": 23, 
                      "SGP_CULTURA": 24, 
                      "SGP_DEPORTE": 25, 
                      "SGP_SALUD": 26,
                      "MUNICIPIO": 27, 
                      "NACION": 28, 
                      "OTROS": 29}
        col_by_idx = {0: "C", 1: "E", 2: "F", 3: "G"}

        from decimal import Decimal
        for yi, y in enumerate(years):
            for ent in ENT_ORDER:
                base_row = row_by_ent[ent]
                col = col_by_idx[yi]
                dest = f"{col}{base_row}"
                val = lookup.get((y, ent), Decimal("0"))
                _write(ws_tecnico, dest, val)

    # ---------------------------
    # 4) Llenar las METAS
    # ---------------------------
    lbl_num = _find_label_anchor(ws, 12, "NÚMERO DE META")
    lbl_nom = _find_label_anchor(ws, 13, "META DE CUATRIENIO")
    lbl_num_col = lbl_num[1] if lbl_num else 2 
    lbl_nom_col = lbl_nom[1] if lbl_nom else 2

    num_val_col = _find_value_anchor_col_for_row(ws, 12, lbl_num_col)
    nom_val_col = _find_value_anchor_col_for_row(ws, 13, lbl_nom_col)

    total = total_metas
    for i in range(1, total + 1):
        dst_start = 12 + (i - 1) * 3
        if i - 1 < len(numero_meta):
            _write(ws, f"{get_column_letter(num_val_col)}{dst_start}", numero_meta[i - 1])
        else:
            _write(ws, f"{get_column_letter(num_val_col)}{dst_start}", "")
        if i - 1 < len(nombre_meta):
            _write(ws, f"{get_column_letter(nom_val_col)}{dst_start + 1}", nombre_meta[i - 1])
        else:
            _write(ws, f"{get_column_letter(nom_val_col)}{dst_start + 1}", "")

    # ---------------------------
    # 4.5) Estructura financiera 
    # ---------------------------
    ef_rows = data.get("estructura_financiera", [])
    if ef_rows:
        years = sorted({row.get("anio") for row in ef_rows if row.get("anio") is not None})[:4]
        while len(years) < 4:
            years.append(None)

        header_cols = ["C", "E", "F", "G"]
        for yi, col in enumerate(header_cols):
            _write(ws, f"{col}{17 + shift}", years[yi] if years[yi] is not None else "")

        ENT_ORDER = ["PROPIOS", "SGP_LIBRE_INVERSION", "SGP_LIBRE_DESTINACION", "SGP_APSB", "SGP_EDUCACION", "SGP_ALIMENTACION_ESCOLAR", "SGP_CULTURA", "SGP_DEPORTE", "SGP_SALUD", "MUNICIPIO", "NACION", "OTROS"]
        row_by_ent = {"PROPIOS": 19, 
                      "SGP_LIBRE_INVERSION": 20, 
                      "SGP_LIBRE_DESTINACION": 21, 
                      "SGP_APSB": 22, 
                      "SGP_EDUCACION": 23, 
                      "SGP_ALIMENTACION_ESCOLAR": 24, 
                      "SGP_CULTURA": 25, 
                      "SGP_DEPORTE": 26, 
                      "SGP_SALUD": 27,
                      "MUNICIPIO": 28, 
                      "NACION": 29, 
                      "OTROS": 30}
        col_by_idx = {0: "C", 1: "E", 2: "F", 3: "G"}

        from decimal import Decimal
        lookup = {}
        for r in ef_rows:
            anio = r.get("anio")
            ent = (r.get("entidad") or "").strip().upper()
            valor = r.get("valor")
            if ent in row_by_ent:
                lookup[(anio, ent)] = valor if valor is not None else Decimal("0")

        for yi, y in enumerate(years):
            for ent in ENT_ORDER:
                base_row = row_by_ent[ent]
                col = col_by_idx[yi]
                dest = f"{col}{base_row + shift}"
                val = lookup.get((y, ent), Decimal("0"))
                _write(ws, dest, val)

    if ef_rows:
        header_cols = ["C", "E", "F", "G"]
        for yi, col in enumerate(header_cols):
            _write(ws_tecnico, f"{col}{16}", years[yi] if years[yi] is not None else "")

        ENT_ORDER = ["PROPIOS", "SGP_LIBRE_INVERSION", "SGP_LIBRE_DESTINACION", "SGP_APSB", "SGP_EDUCACION", "SGP_ALIMENTACION_ESCOLAR", "SGP_CULTURA", "SGP_DEPORTE", "SGP_SALUD", "MUNICIPIO", "NACION", "OTROS"]
        row_by_ent = {"PROPIOS": 18, 
                      "SGP_LIBRE_INVERSION": 19, 
                      "SGP_LIBRE_DESTINACION": 20, 
                      "SGP_APSB": 21, 
                      "SGP_EDUCACION": 22, 
                      "SGP_ALIMENTACION_ESCOLAR": 23, 
                      "SGP_CULTURA": 24, 
                      "SGP_DEPORTE": 25, 
                      "SGP_SALUD": 26,
                      "MUNICIPIO": 27, 
                      "NACION": 28, 
                      "OTROS": 29}
        col_by_idx = {0: "C", 1: "E", 2: "F", 3: "G"}

        from decimal import Decimal
        for yi, y in enumerate(years):
            for ent in ENT_ORDER:
                base_row = row_by_ent[ent]
                col = col_by_idx[yi]
                dest = f"{col}{base_row}"
                val = lookup.get((y, ent), Decimal("0"))
                _write(ws_tecnico, dest, val)

    # ---------------------------
    # 4.9) Llenar las METAS Concepto Tecnico General
    # ---------------------------
    if total_metas > 1:
        extra2 = total_metas - 1
        _move_down_from_row(ws_tecnico, start_row=14, row_off=extra2 * 3)
        merges_base_tecnico = _get_block_merges(ws_tecnico, 11, 13)
        for i in range(2, total_metas + 1):
            dst_start2 = 11 + (i - 1) * 3
            _clone_block_styles_merges(ws_tecnico, 11, 13, dst_start2, merges_cache=merges_base_tecnico)
            lbl_num2 = _find_label_anchor(ws_tecnico, 11, "NÚMERO DE META")
            lbl_nom2 = _find_label_anchor(ws_tecnico, 12, "META DE CUATRIENIO")
            if lbl_num2:
                _write(ws_tecnico, f"{get_column_letter(lbl_num2[1])}{dst_start2}", "NÚMERO DE META")
            if lbl_nom2:
                _write(ws_tecnico, f"{get_column_letter(lbl_nom2[1])}{dst_start2+1}", "META DE CUATRIENIO")

    lbl_num2 = _find_label_anchor(ws_tecnico, 11, "NÚMERO DE META")
    lbl_nom2 = _find_label_anchor(ws_tecnico, 12, "META DE CUATRIENIO")
    lbl_num2_col = lbl_num2[1] if lbl_num2 else 2
    lbl_nom2_col = lbl_nom2[1] if lbl_nom2 else 2
    num_val2_col = _find_value_anchor_col_for_row(ws_tecnico, 11, lbl_num2_col)
    nom_val2_col = _find_value_anchor_col_for_row(ws_tecnico, 12, lbl_nom2_col)

    for i in range(1, total_metas + 1):
        dst_start2 = 11 + (i - 1) * 3
        _write(ws_tecnico, f"{get_column_letter(num_val2_col)}{dst_start2}",
               numero_meta[i - 1] if i - 1 < len(numero_meta) else "")
        _write(ws_tecnico, f"{get_column_letter(nom_val2_col)}{dst_start2 + 1}",
               nombre_meta[i - 1] if i - 1 < len(nombre_meta) else "")

    # ---------------------------
    # 5) Guardar
    # ---------------------------
    out_dir = output_dir or base_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    n = force_index if force_index is not None else _next_sequential_index(out_dir)
    out_path = out_dir / OUTPUT_CONCEPTO.format(n)

    wb.save(str(out_path))
    return out_path

def fill_cadena_valor(base_dir: Path, data: dict, force_index: Optional[int] = None, output_dir: Optional[Path] = None) -> Path:
    template = base_dir / TEMPLATE_CADENA
    wb = load_workbook(str(template))
    ws = wb.active

    _write(ws, "B2", data.get("nombre_proyecto", ""))
    _write(ws, "N2", data.get("cod_id_mga", ""))
    _write(ws, "A21", data.get("fecha_actual", ""))

    out_dir = output_dir or base_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    n = force_index if force_index is not None else _next_sequential_index(out_dir)
    out_path = out_dir / OUTPUT_CADENA.format(n)

    wb.save(str(out_path))
    return out_path

def fill_viabilidad_dependencias(base_dir: Path, data: dict, force_index: Optional[int] = None, output_dir: Optional[Path] = None) -> Path:
    template = base_dir / TEMPLATE_VIABILIDAD
    wb = load_workbook(str(template))
    ws = wb.active

    _write(ws, "G3", data.get("dependencia", ""))
    _write(ws, "G5", data.get("nombre_proyecto", ""))
    _write(ws, "G6", data.get("cod_id_mga", ""))

    anios = data.get("anios", [])
    for i, c in enumerate(["D9", "G9", "J9", "M9"]):
        if i < len(anios):
            _write(ws, c, anios[i])

    ent_row = {
        "PROPIOS": 11, "SGP_LIBRE_INVERSION": 12, "SGP_LIBRE_DESTINACION": 13, "SGP_APSB": 14,
        "SGP_EDUCACION": 15, "SGP_ALIMENTACION_ESCOLAR": 16, "SGP_CULTURA": 17, "SGP_DEPORTE": 18,
        "SGP_SALUD": 19, "MUNICIPIO": 20, "NACION": 21, "OTROS": 22,
    }
    col_map = {0: "D", 1: "G", 2: "J", 3: "M"}
    lookup = data.get("estructura_financiera", {})

    for yi, anio in enumerate(anios):
        for ent, row in ent_row.items():
            val = lookup.get((anio, ent), 0)
            _write(ws, f"{col_map[yi]}{row}", val)
    
    ids_sel = set(int(x) for x in data.get("viabilidades_ids", data.get("viabilidades", [])) if x)
    id_to_row = data.get("viabilidad_id_to_row") or {}
    for i in range(1, 7):
        id_to_row.setdefault(i, 32 + i)
    for r in range(33, 39):
        _write(ws, f"S{r}", "NO")
    for vid in ids_sel:
        r = id_to_row.get(vid)
        if r and 33 <= r <= 38:
            _write(ws, f"S{r}", "SI")

    funcs = data.get("funcionarios", {})
    for itv, f in funcs.items():
        if itv == 1:
            _write(ws, "B41", f"Funcionario que certifica viabilidad técnica:\nNombre: {f.nombre}\nCargo: {f.cargo}")
        elif itv == 2:
            _write(ws, "B42", f"Funcionario que certifica viabilidad jurídica:\nNombre: {f.nombre}\nCargo: {f.cargo}")
        elif itv == 3:
            _write(ws, "B43", f"Funcionario que certifica viabilidad financiera:\nNombre: {f.nombre}\nCargo: {f.cargo}")

    _write(ws, "E48", data.get("nombre_secretario", ""))
    _write(ws, "E49", data.get("dependencia", ""))
    _write(ws, "E50", data.get("fecha_actual", ""))

    out_dir = output_dir or base_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    n = force_index if force_index is not None else _next_sequential_index(out_dir)
    out_path = out_dir / OUTPUT_VIABILIDAD.format(n)

    wb.save(str(out_path))
    return out_path

def _normaliza_resp_str(v: str) -> str:
    s = (v or "").strip().upper()
    if s in {"SI", "SÍ"}:
        return "SI"
    if s == "NO":
        return "NO"
    if s in {"N/A", "NA"}:
        return "N/A"
    return s

def _to_respuesta(v) -> str:
    if isinstance(v, dict):
        if "RESPUESTA" in v:
            return _normaliza_resp_str(v["RESPUESTA"])
        if "respuesta" in v:
            return _normaliza_resp_str(v["respuesta"])
        for k in ("valor", "value", "resp", "answer"):
            if k in v:
                return _normaliza_resp_str(v[k])
        return ""
    if isinstance(v, bool):
        return "SI" if v else "NO"
    if isinstance(v, str):
        return _normaliza_resp_str(v)
    return ""